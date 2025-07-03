
# -*- coding: utf-8 -*-
"""
Created on Tue Jul 1 16:39:24 2025
@author: amin.sahabi
"""
import sys
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter
import os
import re
import copy

plantilla_file = sys.argv[1]
specs_file = "SPECS CLIENTE base.xlsx"

df_gas = pd.read_excel(plantilla_file, sheet_name="GAS", engine="openpyxl", header=4)
df_lista = pd.read_excel(plantilla_file, sheet_name="Lista", engine="openpyxl")

peaje_to_distribuidora = dict(zip(df_lista["Peaje Gas"], df_lista["Gestión ATR"]))

wb = load_workbook(specs_file)
ws = wb["SPECS"]
ws_price_ka = wb["Price KA"]

start_col = 4
row_counterparty = 4
row_cif = 5
row_cups = 13
row_tariff = 14
row_qd = 15
row_distribuidora = 16
row_start_date = 7
col_start_date = 4  # D
col_end_date = 5    # E
row_consumo_start = 20
row_consumo_end = 55
row_total_volume = 56
row_avg_volume = 57

unique_clients = df_gas["Razón Social"].dropna().unique()
nombre_cliente = unique_clients[0] if len(unique_clients) == 1 else "CLIENTE"
output_file = os.path.join("salida", f"SPECS {nombre_cliente} (X).xlsx")

if len(unique_clients) == 1:
    ws.cell(row=row_counterparty, column=start_col).value = unique_clients[0]
if len(df_gas["CIF"].dropna().unique()) == 1:
    ws.cell(row=row_cif, column=start_col).value = df_gas["CIF"].dropna().unique()[0]

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

original_widths = {}
for col in range(4, 8):
    letter = get_column_letter(col)
    original_widths[col] = ws.column_dimensions[letter].width

cups_index = 0
for _, row in df_gas.iterrows():
    if pd.isna(row["CUPS"]):
        continue

    col = start_col + cups_index

    if len(unique_clients) > 1:
        ws.cell(row=row_counterparty, column=col).value = row["Razón Social"]
    if len(df_gas["CIF"].dropna().unique()) > 1:
        ws.cell(row=row_cif, column=col).value = row["CIF"]

    ws.cell(row=row_cups, column=col).value = row["CUPS"]
    ws.cell(row=row_tariff, column=col).value = row["Peaje"]

    qd_value = row["Qd contratada (kWh/día)"]
    if isinstance(qd_value, (int, float)):
        ws.cell(row=row_qd, column=col).value = qd_value
    else:
        ws.cell(row=row_qd, column=col).value = None

    distribuidora = peaje_to_distribuidora.get(row["Peaje"], "")
    ws.cell(row=row_distribuidora, column=col).value = distribuidora

    if cups_index == 0:
        ws.cell(row=row_start_date, column=col_start_date).value = row["Fecha inicio"]
        ws.cell(row=row_start_date, column=col_end_date).value = row["Fecha fin"]

    for i in range(12):
        ws.cell(row=row_consumo_start + i, column=col).value = row.iloc[14 + i]

    for r in list(range(row_cups, row_distribuidora + 1)) + list(range(row_consumo_start, row_consumo_start + 12)):
        cell = ws.cell(row=r, column=col)
        if cell.value not in [None, ""]:
            cell.border = thin_border

    if col > 7 and 4 in original_widths:
        ws.column_dimensions[get_column_letter(col)].width = original_widths[4]

    cups_index += 1

last_col_letter = get_column_letter(start_col + cups_index - 1)
for row in range(row_consumo_start, row_consumo_end + 1):
    formula = f"=SUM(D{row}:{last_col_letter}{row})"
    ws.cell(row=row, column=2).value = formula

ws.cell(row=row_total_volume, column=3).value = f"=SUM(D{row_total_volume}:{last_col_letter}{row_total_volume})"
ws.cell(row=row_avg_volume, column=3).value = f"=SUM(D{row_avg_volume}:{last_col_letter}{row_avg_volume})"

# Añadir fórmulas correctas en cada columna de CUPS
for col in range(start_col, start_col + cups_index):
    col_letter = get_column_letter(col)
    ws[f"{col_letter}{row_total_volume}"] = f"=SUM({col_letter}{row_consumo_start}:{col_letter}{row_consumo_end})"
    ws[f"{col_letter}{row_avg_volume}"] = f"=IFERROR({col_letter}{row_total_volume}/COUNTA($C${row_consumo_start}:$C${row_consumo_end})*12,0)"

forma_pago_values = df_gas["Forma de pago"].dropna().astype(str).values
if len(forma_pago_values) > 0:
    match = re.search(r"\d+", forma_pago_values[0])
    if match:
        ws_price_ka.cell(row=7, column=13).value = int(match.group())  # M7

# === NUEVO BLOQUE: Añadir columnas en 'Price KA' si hay más de 4 CUPS ===
if "Price KA" in wb.sheetnames:
    ws_price_ka = wb["Price KA"]
    base_cups = 4
    extra_cups = max(0, cups_index - base_cups)

    if extra_cups > 0:
        base_start_col = 4  # Columna D
        base_end_col = 7    # Columna G
        insert_at = base_end_col + 1  # Insertar después de G

        # Insertar columnas vacías
        for _ in range(extra_cups):
            ws_price_ka.insert_cols(insert_at)

        # Copiar contenido y formato desde columnas D-G
        for i in range(extra_cups):
            src_col = base_start_col + (i % (base_end_col - base_start_col + 1))
            tgt_col = insert_at + i
            for row in ws_price_ka.iter_rows(min_row=1, max_row=ws_price_ka.max_row):
                src_cell = row[src_col - 1]
                tgt_cell = ws_price_ka.cell(row=src_cell.row, column=tgt_col)
                tgt_cell.value = src_cell.value
                if src_cell.has_style:
                    tgt_cell.font = copy.copy(src_cell.font)
                    tgt_cell.border = copy.copy(src_cell.border)
                    tgt_cell.fill = copy.copy(src_cell.fill)
                    tgt_cell.number_format = copy.copy(src_cell.number_format)
                    tgt_cell.protection = copy.copy(src_cell.protection)
                    tgt_cell.alignment = copy.copy(src_cell.alignment)
            # Copiar ancho de columna
            src_letter = get_column_letter(src_col)
            tgt_letter = get_column_letter(tgt_col)
            ws_price_ka.column_dimensions[tgt_letter].width = ws_price_ka.column_dimensions[src_letter].width
            
            
wb.save(output_file)
print(f"Archivo generado: {output_file}")

